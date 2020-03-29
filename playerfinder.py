from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.formatting.rule import ColorScale, FormatObject, CellIsRule, ColorScaleRule, FormulaRule, Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from decimal import *
import pandas
import json
import decimal

wb = Workbook()
ws = wb.active
redFill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
getcontext()
Context(prec=28, rounding=ROUND_HALF_EVEN, Emin=-999999, Emax=999999,
        capitals=1, clamp=0, flags=[], traps=[Overflow, DivisionByZero,
                                              InvalidOperation])
getcontext().prec = 4

with open('player_info.JSON', 'r') as myfile:
    data = myfile.read()

versusArray = [["TSM", "EG"], ["C9", "CLG"], ["DIG", "GG"], ["TL", "100"], ["FLY", "IMT"]]
obj = json.loads(data)

def versusArrayFunc():

    for i in range(5):
        versusArray.append([input("Who is the first team?\n"), input("Who are they going against?\n")])

def initial_formatting():
    ws.row_dimensions[1].height = 25.5
    ws.row_dimensions[2].height = 21.75
    ws.row_dimensions[3].height = 27

    ws.column_dimensions['K'].width = .5
    ws.column_dimensions['I'].width = .5
    ws.column_dimensions['O'].width = 5.86
    ws.column_dimensions['P'].width = 5.86
    ws.column_dimensions['Q'].width = 5.86
    ws.column_dimensions['R'].width = 5.86
    ws.column_dimensions['S'].width = 5.86
    ws.column_dimensions['T'].width = 5.86

    ws.merge_cells('A1:J2')
    ws.merge_cells('L1:V1')
    ws.merge_cells('L2:N2')
    ws.merge_cells('O2:P2')
    ws.merge_cells('Q2:R2')
    ws.merge_cells('S2:T2')

    ws.cell(row=3, column=1).value = "Name"
    ws.cell(row=3, column=2).value = "Team"
    ws.cell(row=3, column=3).value = "POS"
    ws.cell(row=3, column=4).value = "Opp "
    ws.cell(row=3, column=5).value = "Proj "
    ws.cell(row=3, column=6).value = "GD2  "
    ws.cell(row=3, column=7).value = "MP2  "
    ws.cell(row=3, column=8).value = "AvG  "
    ws.cell(row=3, column=9).value = "AMP"
    ws.cell(row=3, column=10).value = "GPM"
    ws.cell(row=1, column=12).value = "Recent Form"
    ws.cell(row=2, column=12).value = "Recent High"
    ws.cell(row=2, column=15).value = "L6"
    ws.cell(row=2, column=17).value = "L4"
    ws.cell(row=2, column=19).value = "L2"
    ws.cell(row=2, column=19).value = "EVAL"
    ws.cell(row=3, column=15).value = "L6Av"
    ws.cell(row=3, column=16).value = "L6:Val"
    ws.cell(row=3, column=17).value = "L4Av"
    ws.cell(row=3, column=18).value = "L4:Val"
    ws.cell(row=3, column=19).value = "L2Av"
    ws.cell(row=3, column=20).value = "L2:Val"
    ws.cell(row=3, column=21).value = "GD1"
    ws.cell(row=3, column=22).value = "MP1"
    ws.cell(row=3, column=23).value = "GD2"
    ws.cell(row=3, column=24).value = "MP2"
    ws.cell(row=3, column=25).value = "AvG"
    ws.cell(row=3, column=26).value = "AvMP"
    ws.cell(row=3, column=27).value = "GPM"

    ws['L3'].alignment = Alignment(wrap_text=True)
    ws.cell(row=3, column=12).value = "Last\nGame"
    ws['M3'].alignment = Alignment(wrap_text=True)
    ws.cell(row=3, column=13).value = "L10\nHigh"
    ws.cell(row=3, column=14).value = "VAL"

    ws.cell(row=3, column=1).font = Font(size=14, bold=True)
    ws.cell(row=3, column=2).font = Font(size=12)
    ws.cell(row=3, column=3).font = Font(size=12)
    ws.cell(row=3, column=4).font = Font(size=12)
    ws.cell(row=3, column=5).font = Font(size=12)
    ws.cell(row=3, column=6).font = Font(size=12)
    ws.cell(row=3, column=7).font = Font(size=12)
    ws.cell(row=3, column=8).font = Font(size=12)
    ws.cell(row=3, column=12).font = Font(size=11)
    ws.cell(row=3, column=13).font = Font(size=11)
    ws.cell(row=3, column=14).font = Font(size=11)
    ws.cell(row=2, column=12).font = Font(size=13, bold=True)
    ws.cell(row=2, column=15).font = Font(size=13, bold=True)
    ws.cell(row=2, column=17).font = Font(size=13, bold=True)
    ws.cell(row=2, column=19).font = Font(size=13, bold=True)

    ws.cell(row=2, column=12).alignment = Alignment(horizontal='center')
    ws.cell(row=2, column=15).alignment = Alignment(horizontal='center')
    ws.cell(row=2, column=17).alignment = Alignment(horizontal='center')
    ws.cell(row=2, column=19).alignment = Alignment(horizontal='center')

    # Recent Form L1 Formatting
    ws.cell(row=1, column=12).alignment = Alignment(horizontal='center')
    ws.cell(row=1, column=12).font = Font(size=15, bold=True)

def post_formatting():
    column_widths = []
    for row in ws.iter_rows(3, 55):
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                try:
                    column_widths.append(len(cell.value))
                except TypeError:
                    print("Type Error")

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width + 1

    ws.conditional_formatting.add('D4:D54',
                                  ColorScaleRule(start_type='percentile', start_value=10, start_color='ea7d7d',
                                                 mid_type='percentile', mid_value=50, mid_color='C0C0C0',
                                                 end_type='percentile', end_value=90, end_color='9de7b1'))

    ws.conditional_formatting.add('F4:F54',
                                  ColorScaleRule(start_type='percentile', start_value=10, start_color='ea7d7d',
                                                 mid_type='percentile', mid_value=50, mid_color='C0C0C0',
                                                 end_type='percentile', end_value=90, end_color='9de7b1'))
    ws.conditional_formatting.add('H4:H54',
                                  ColorScaleRule(start_type='percentile', start_value=10, start_color='AA0000',
                                                 mid_type='percentile', mid_value=50, mid_color='C0C0C0',
                                                 end_type='percentile', end_value=90, end_color='00AA00'))

    ws.conditional_formatting.add('J4:J54',
                                  ColorScaleRule(start_type='percentile', start_value=10, start_color='ffffff',
                                                 mid_type='percentile', mid_value=50, mid_color='ffe3a3',
                                                 end_type='percentile', end_value=90, end_color='ffc73b'))

    red_text = Font(color="ffffff")
    red_fill = PatternFill(bgColor="131313")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("TSM",B1)))']
    ws.conditional_formatting.add('B1:F61', rule)

    red_text = Font(color="ffffff")
    red_fill = PatternFill(bgColor="149fda")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("C9",B1)))']
    ws.conditional_formatting.add('B1:F61', rule)

    red_text = Font(color="131313")
    red_fill = PatternFill(bgColor="FFFF66")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("DIG",B1)))']
    ws.conditional_formatting.add('B1:F61', rule)

    red_text = Font(color="ff1d1d")
    red_fill = PatternFill(bgColor="131313")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("100T",B1)))']
    ws.conditional_formatting.add('B1:F61', rule)

    red_text = Font(color="ffffff")
    red_fill = PatternFill(bgColor="092f7e")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("TL",B1)))']
    ws.conditional_formatting.add('B1:F61', rule)

    red_text = Font(color="3399ff")
    red_fill = PatternFill(bgColor="131313")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("CLG",B1)))']
    ws.conditional_formatting.add('B1:F61', rule)

    red_text = Font(color="ffffff")
    red_fill = PatternFill(bgColor="001a33")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("EG",B1)))']
    ws.conditional_formatting.add('B1:F61', rule)

    red_text = Font(color="004d00")
    red_fill = PatternFill(bgColor="e6b800")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("FLY",B1)))']
    ws.conditional_formatting.add('B1:F61', rule)

    red_text = Font(color="33cccc")
    red_fill = PatternFill(bgColor="0a2929")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("IMT",B1)))']
    ws.conditional_formatting.add('B1:F61', rule)

    red_text = Font(color="ffcc33")
    red_fill = PatternFill(bgColor="131313")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("GG",B1)))']
    ws.conditional_formatting.add('B1:F61', rule)


    for rows in ws.iter_rows(min_row=1, max_row=54, min_col=11, max_col=11):
        for cell in rows:
            cell.fill = PatternFill(bgColor="FFC7CE", fill_type="solid")

    ws.column_dimensions['K'].width = .5
def main():
    initial_formatting()
    j = 0
    k = 0
    row = 4
    exists = False
    first_column = ws['A']
    avgGold = 0

    for week in obj:
        for day in obj[week]:
            for team in obj[week][day]:
                for player in obj[week][day][team]:
                    first_column = ws['A']
                    try:
                        for x in range(len(first_column)):  # Checks if player exists
                            if first_column[x].value == player:
                                ws.cell(row=x + 1, column=23).value = float(obj[week][day][team][player]['gold'])
                                ws.cell(row=x + 1, column=25).value = Decimal(
                                    (ws.cell(row=x + 1, column=23).value + ws.cell(row=x + 1, column=21).value)) / Decimal(2)
                                ws.cell(row=x + 1, column=24).value = obj[week][day][team]['duration']
                                minutes1 = Decimal(ws.cell(row=x + 1, column=22).value.split(":")[0])
                                seconds1 = Decimal(ws.cell(row=x + 1, column=22).value.split(":")[1])
                                minutes2 = Decimal(ws.cell(row=x + 1, column=24).value.split(":")[0])
                                seconds2 = Decimal(ws.cell(row=x + 1, column=24).value.split(":")[1])
                                avgMinutes = (minutes1 + minutes2) / Decimal(2)
                                avgSeconds = (seconds1 + seconds2) / Decimal(2)
                                ws.cell(row=x + 1, column=26).value = str(int(avgMinutes)) + ":" + str(int(avgSeconds))

                                (m, s) = ws.cell(row=x + 1, column=26).value.split(':')
                                result = int(m) * 60 + int(s)
                                ws.cell(row=x + 1, column=27).value = Decimal(1000) * Decimal(ws.cell(row=x + 1, column=25).value) / (Decimal(result) / Decimal(60))
                                exists = True
                        if exists is False:
                            ws.cell(row=row, column=3).value = obj[week][day][team][player]['position'] + " "
                            ws.cell(row=row, column=1).value = player
                            if team == '100':
                                ws.cell(row=row, column=2).value = "100T"
                            else:
                                ws.cell(row=row, column=2).value = team
                            ws.cell(row=row, column=21).value = float(obj[week][day][team][player]['gold'])
                            ws.cell(row=row, column=22).value = obj[week][day][team]['duration']
                            for i in range(len(versusArray)):
                                if versusArray[i][0] == team:
                                    if versusArray[i][1] == '100':
                                        ws.cell(row=row, column=4).value = "100T"
                                    else:
                                        ws.cell(row=row, column=4).value = versusArray[i][1]
                                elif versusArray[i][1] == team:
                                    if versusArray[i][0] == '100':
                                        ws.cell(row=row, column=4).value = '100T'
                                    else:
                                        ws.cell(row=row, column=4).value = versusArray[i][0]
                            print(obj[week][day][team]['duration'])
                            row += 1
                    except TypeError:
                        pass

                    exists = False
    #versusArrayFunc()
    #print(versusArray)

main()
post_formatting()
wb.save("lwt example.xlsx")
