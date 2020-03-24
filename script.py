from urllib.request import urlopen as uReq
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.formatting.rule import ColorScale, FormatObject, CellIsRule, ColorScaleRule
from decimal import *
import pandas
import json
from selenium import webdriver  # Import module
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
import time  # Waiting function
import re
import copy

# This enables floats to be manipulated by the decimal library
wb = Workbook()
ws = wb.active

browser = webdriver.Chrome()
wait = WebDriverWait(browser, 10)
# Define URL
uClient = uReq("https://lol.gamepedia.com/LCS/2020_Season/Spring_Season/Scoreboards")
page_html = uClient.read()
page_soup = BeautifulSoup(page_html, "html.parser")

list = page_soup.findAll("div", attrs={"class": "inline-content"})
weekLinks = page_soup.find_all('div', {'class': 'tabheader-top'})[2].find_all('a', href=re.compile("Week"))

matchHistoryPage = []
playerLength = 0
minGold = 0
day1Dict = {}
day2Dict = {}
day3Dict = {}
weekDict = {'DAY1': day1Dict, 'DAY2': day2Dict, 'DAY3': day3Dict}
seasonDick = {}
positionArray = ['TOP', 'JUNG', 'MID', 'ADC', 'SUP']


def in_list(item, L):
    if item is None:
        return -1
    else:
        for x in L:
            if item in x:
                return L.index(x)
        return -1


def add_k(sheet):
    for row in sheet.iter_rows(min_row=2, min_col=2):
        for cell in row:
            if cell.value is not None:
                if float(cell.value):
                    cell.value = str(cell.value) + "k "


def format_color_cells(sheet):
    sheet.conditional_formatting.add('B1:B52',
                                     ColorScaleRule(start_type='percentile', start_value=10, start_color='ea7d7d',
                                                    mid_type='percentile', mid_value=50, mid_color='C0C0C0',
                                                    end_type='percentile', end_value=90, end_color='9de7b1'))

    sheet.conditional_formatting.add('D1:D52',
                                     ColorScaleRule(start_type='percentile', start_value=10, start_color='ea7d7d',
                                                    mid_type='percentile', mid_value=50, mid_color='C0C0C0',
                                                    end_type='percentile', end_value=90, end_color='9de7b1'))
    sheet.conditional_formatting.add('F1:F52',
                                     ColorScaleRule(start_type='percentile', start_value=10, start_color='AA0000',
                                                    mid_type='percentile', mid_value=50, mid_color='C0C0C0',
                                                    end_type='percentile', end_value=90, end_color='00AA00'))


def is_loaded(path):
    # path example: //a[@class="login-button btn-large btn-large-primary"]
    try:
        e = browser.find_element(By.XPATH, path)
    except NoSuchElementException:
        print("not loaded yet")
        e = None
    while e is None:
        try:
            e = browser.find_element(By.XPATH, path)
        except NoSuchElementException:
            print("not loaded yet")
            e = None
    return e


def get_links():
    for t in list:  # finds links for week 1
        matchLink = t.find_all("div", {"class": "sb-datetime-mh"})
        for div in matchLink:
            i = div.find('a')
            matchHistoryPage.append(i['href'])

    for week in weekLinks:  # finds links for rest of weeks
        print(week['href'])
        uClient = uReq("https://lol.gamepedia.com" + week['href'])
        page_html = uClient.read()
        page_soup = BeautifulSoup(page_html, "html.parser")
        games = page_soup.findAll("div", attrs={"class": "inline-content"})
        for game in games:
            matchLink = game.find_all("div", {"class": "sb-datetime-mh"})
            for div in matchLink:
                i = div.find('a')
                matchHistoryPage.append(i['href'])


def main():
    get_links()
    browser.get('https://matchhistory.na.leagueoflegends.com/en/#page/landing-page')
    is_loaded('//a[@class="login-button btn-large btn-large-primary"]').click()
    is_loaded('//input[@name="username"]').send_keys('dryrhino4419')
    is_loaded('//input[@name="password"]').send_keys('gabythebaby1')
    time.sleep(7)
    browser.find_element_by_name("password").send_keys(Keys.ENTER)
    time.sleep(7)

    masterI = 0
    for i in range(len(matchHistoryPage)):  # loops through each match history page
        browser.get(matchHistoryPage[i])
        browser.refresh()
        time.sleep(3)
        element = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@class="classic details"]')))
        page_html = element.get_attribute('innerHTML')
        page_soup = BeautifulSoup(page_html, "html.parser")

        playersRow = page_soup.findAll('div', {"class": "classic player"})
        teamDict = {}
        playerDict = {}
        j = 0
        k = 0

        statTable = page_soup.find('table', {"class": "table table-bordered"})
        minionsKilledRow = statTable.contents[1].contents[32]
        neutralMinionsKilledRow = statTable.contents[1].contents[33]
        neutralKilledInTEAMJRow = statTable.contents[1].contents[34]
        neutralKilledInENMJRow = statTable.contents[1].contents[35]
        largestKillingSpreeRow = statTable.contents[1].contents[3]
        duration = page_soup.find('span', {"class": "map-header-duration"})

        for player in playersRow:
            statDick = {}
            if j == 5:
                playerDict = {}
                statDick = {}
                k = 0

            statDick.update({"index": k})
            statDick.update({"position": positionArray[k]})

            statDick.update(
                {"champion": player.find('div', {"class": 'champion-icon binding'}).contents[0].attrs['data-rg-id']})

            if j > 4:
                teamName = player.find('div', {"class": "champion-nameplate-name"}).get_text().split(" ", 3)[1]
                statDick.update({"laneVS": playersRow[j - 5].find('div', {
                    "class": "champion-nameplate-name"}).get_text().split(" ", 3)[2]})
                statDick.update({"oppCS": minionsKilledRow.contents[j - 4].get_text()})
            else:
                statDick.update({"laneVS": playersRow[j + 5].find('div', {
                    "class": "champion-nameplate-name"}).get_text().split(" ", 3)[2]})
                statDick.update({"oppCS": minionsKilledRow.contents[j + 6].get_text()})

            statDick.update({"gold": player.find('div', {"class": "gold-col gold"}).text[:-1]})
            statDick.update({"kills": player.find('div', {"class": "kda-kda"}).get_text().split("/", 3)[0]})
            statDick.update({"deaths": player.find('div', {"class": "kda-kda"}).get_text().split("/", 3)[1]})
            statDick.update({"assists": player.find('div', {"class": "kda-kda"}).get_text().split("/", 3)[2]})
            statDick.update({"minions_killed": minionsKilledRow.contents[j + 1].get_text()})
            statDick.update({"neutral_minions_killed": neutralMinionsKilledRow.contents[j + 1].get_text()})
            statDick.update({"neutral_minions_killed_team_jungle": neutralKilledInTEAMJRow.contents[j + 1].get_text()})
            statDick.update({"neutral_minions_killed_enemy_jungle": neutralKilledInENMJRow.contents[j + 1].get_text()})
            statDick.update({"largest_killing_spree": largestKillingSpreeRow.contents[j + 1].get_text()})

            playerDict.update({'duration': duration.get_text()})
            playerDict.update({player.find('div', {"class": "champion-nameplate-name"}).get_text().split(" ", 3)[2]: statDick})

            k += 1

            teamDict.update(
                {player.find('div', {"class": "champion-nameplate-name"}).get_text().split(" ", 3)[1]: playerDict})

            j += 1

        if masterI < 4:
            weekDict['DAY1'].update(teamDict)
        elif masterI == 4 or masterI < 8:
            weekDict['DAY2'].update(teamDict)
        else:
            weekDict['DAY3'].update(teamDict)

        if masterI == 9:
            seasonDick.update({'WEEK' + str(-(-(i+1) // 10)): copy.deepcopy(weekDict)})
            day1Dict.clear()
            day2Dict.clear()
            day3Dict.clear()
            weekDict.update({'DAY1': day1Dict, 'DAY2': day2Dict, 'DAY3': day3Dict})
            print(-(-(i+1) // 10))
            masterI = 0
        else:
            masterI += 1

    with open('player_info.JSON', 'w', encoding='utf-8') as f:
        json.dump(seasonDick, f, ensure_ascii=False, indent=4)

    add_k(ws)


main()

'''
        allTheStats = {}

        for x in range(2, len(statTable.contents[1].contents)):
            row = statTable.contents[1].contents[x]
            tet = row.attrs['class'][0].strip()
            v = 'view'
            if tet != v.strip():
                statlist = []
                for y in range(1, len(row.contents)):
                    statlist.append(row.contents[y].get_text())
                allTheStats.update({row.contents[0].get_text(): statlist})



'''
